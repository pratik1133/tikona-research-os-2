"""Offline driver — exercises everything in financial_model_v5.py EXCEPT the
Claude / Screener / yfinance / Supabase network calls. Loads the saved
GRAVITA_screener.xlsx + GRAVITA_model.json, runs normalize → validate →
build_model, then inspects the output workbook.

Run from repo root:
    .venv/Scripts/python.exe scripts/offline_test_build_model.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

# Force UTF-8 stdout so Windows console doesn't choke on unicode in log output.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

# Stub out the heavy deps that the FM script imports at top-level but
# the build_model code path doesn't actually exercise.
import types
for mod_name in ("anthropic", "yfinance", "cloudscraper"):
    if mod_name not in sys.modules:
        stub = types.ModuleType(mod_name)
        sys.modules[mod_name] = stub

# bs4 is used in fetch_market_data which we don't call, but it's imported
# at module top — provide minimal stub if absent.
try:
    import bs4  # noqa: F401
except ModuleNotFoundError:
    stub = types.ModuleType("bs4")
    class _BS:
        def __init__(self, *a, **k): pass
    stub.BeautifulSoup = _BS
    sys.modules["bs4"] = stub

# Add scripts/ to path so the v5 module imports cleanly.
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import financial_model_v5 as fm

SCREENER_XLSX = ROOT / "output" / "GRAVITA_screener.xlsx"
MODEL_JSON    = ROOT / "output" / "GRAVITA_model.json"
OUT_XLSX      = ROOT / "output" / "GRAVITA_model_v52_test.xlsx"
OUT_JSON      = ROOT / "output" / "GRAVITA_model_v52_test.json"

def banner(msg):
    bar = "═" * (len(msg) + 4)
    print(f"\n{bar}\n  {msg}\n{bar}")

def main():
    banner("1. Load screener data + raw model JSON")
    screener = fm.extract_screener_data(str(SCREENER_XLSX))
    print(f"   Screener years: {screener['fiscal_years']}")
    print(f"   CMP: {screener['cmp']}  MCap: {screener['mcap']}")
    raw_model = json.loads(MODEL_JSON.read_text())
    print(f"   Raw JSON: scenario_analysis={'scenario_analysis' in raw_model}, "
          f"catalyst_timeline={'catalyst_timeline' in raw_model}, "
          f"saarthi_dimensions={'saarthi_dimensions' in raw_model.get('thesis', {})}")

    banner("2. normalize_model_output — should synthesize the 3 new canonical fields")
    normalized = fm.normalize_model_output(raw_model, screener)
    print(f"   After normalize: scenario_analysis={'scenario_analysis' in normalized}, "
          f"catalyst_timeline len={len(normalized.get('catalyst_timeline', []))}, "
          f"saarthi_dimensions len={len(normalized.get('thesis', {}).get('saarthi_dimensions', []))}")

    banner("3. validate_model_output — Pydantic schema check")
    validated = fm.validate_model_output(normalized, "GRAVITA")
    print(f"   ✔ schema valid. saarthi_total={validated.thesis.saarthi_total}, "
          f"scenarios prob sum="
          f"{validated.scenario_analysis.bull.probability_pct + validated.scenario_analysis.base.probability_pct + validated.scenario_analysis.bear.probability_pct}")

    # Embed historicals + run deterministic derivation (post v5.3 refactor).
    # This is the single place that produces target_price, upside_pct, projected_pl,
    # valuation fair values, scenario target prices, and weighted_tp. Without it
    # those fields stay None and build_model crashes downstream.
    derived = validated.model_dump()
    fm._embed_pl_series(derived, screener)
    fm.compute_derived_facts(derived, screener)

    # Persist the validated + derived JSON for inspection
    OUT_JSON.write_text(json.dumps(derived, indent=2, default=str))
    print(f"   ↳ wrote {OUT_JSON.name}")
    print(f"   Derived: target_price={derived.get('target_price')} "
          f"upside_pct={derived.get('upside_pct')} "
          f"weighted_tp={derived.get('scenario_analysis', {}).get('weighted_tp')}")

    banner("4. build_model — generate full Excel with all new sheets")
    fm.build_model(str(SCREENER_XLSX), screener, derived, str(OUT_XLSX))
    print(f"   ↳ wrote {OUT_XLSX.name} ({OUT_XLSX.stat().st_size // 1024} KB)")

    banner("5. Inspect output workbook")
    from openpyxl import load_workbook
    wb = load_workbook(str(OUT_XLSX), data_only=False)  # keep formulas as strings
    sheets = wb.sheetnames
    print(f"   Sheets ({len(sheets)}): {sheets}")

    # Required new sheets
    required = ["SAARTHI", "Scenario_Analysis", "Catalyst_Timeline", "Charts"]
    missing = [s for s in required if s not in sheets]
    if missing:
        print(f"   ❌ MISSING new sheets: {missing}")
        return 1
    print(f"   ✔ all 4 new sheets present")

    # SAARTHI sheet formula checks
    print("\n   SAARTHI sheet cells:")
    s = wb["SAARTHI"]
    for cell in ["A4", "B4", "C4", "E5", "C12", "E12", "C13"]:
        print(f"     {cell} = {s[cell].value!r}")

    # Scenario_Analysis formula checks
    print("\n   Scenario_Analysis sheet cells:")
    sa = wb["Scenario_Analysis"]
    for cell in ["B11", "C11", "D11", "B14"]:
        print(f"     {cell} = {sa[cell].value!r}")

    # Cover formula checks
    print("\n   Cover sheet KPI block:")
    cv = wb["Cover"]
    for cell in ["A4", "B4", "B5", "B6", "B7", "B8"]:
        print(f"     {cell} = {cv[cell].value!r}")
    print(f"     A12 = {cv['A12'].value!r}")
    print(f"     E5  = {cv['E5'].value!r}   (SAARTHI dim 1 score display)")
    print(f"     E12 = {cv['E12'].value!r}  (SAARTHI total display)")

    # Valuation formula checks
    print("\n   Valuation sheet:")
    v = wb["Valuation"]
    for cell in ["A3", "B3", "B6", "B7", "B8", "B9", "B10"]:
        print(f"     {cell} = {v[cell].value!r}")

    # Charts sheet — count chart objects
    print("\n   Charts sheet:")
    ch = wb["Charts"]
    n_charts = len(ch._charts)
    print(f"     {n_charts} chart objects embedded")
    if n_charts == 0:
        print("     ❌ no charts on Charts sheet!")
        return 1

    # Catalyst_Timeline sample
    print("\n   Catalyst_Timeline sheet first 3 rows:")
    ct = wb["Catalyst_Timeline"]
    for r in range(5, 8):
        row_vals = [ct.cell(row=r, column=c).value for c in range(1, 5)]
        print(f"     row {r}: {row_vals}")

    banner("ALL CHECKS PASSED")
    print(f"\n   Open {OUT_XLSX.name} in Excel/LibreOffice to verify charts render and formulas recalc.")
    return 0

if __name__ == "__main__":
    sys.exit(main())
