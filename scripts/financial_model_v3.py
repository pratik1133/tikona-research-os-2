"""
Financial Model Generator v4.0
================================
Architecture: 2 LLM calls + deterministic openpyxl builder

Call 1  │ Full research + historical + projections + valuation → JSON
        │ Model: claude-sonnet-4-20250514 + web_search
        │ Max tokens: 16,000

Call 2  │ Investment thesis (SAARTHI) → JSON
        │ Model: claude-sonnet-4-20250514
        │ Max tokens: 1,500

Python  │ Deterministic openpyxl builder → 10-sheet Excel
        │ No LLM. Balance checks enforced in code. Always executes.

Cost: ~$0.65/run  |  Quality: institutional-grade  |  Reliability: high
"""

import sys, os, re, json, time, traceback
from typing import Any, Optional

import anthropic

# ── Anthropic ──────────────────────────────────────────────────────────────
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
SONNET = "claude-sonnet-4-20250514"
WEB_SEARCH_TOOL = [{"type": "web_search_20250305", "name": "web_search"}]


# ══════════════════════════════════════════════════════════════════════════════
# FREE DATA FETCHERS
# ══════════════════════════════════════════════════════════════════════════════

def fetch_yfinance_data(ticker_ns: str) -> dict:
    try:
        import yfinance as yf
        stock = yf.Ticker(ticker_ns)
        info = stock.info

        def df_to_dict(df):
            if df is None or df.empty:
                return {}
            df.columns = [str(c.date()) if hasattr(c, 'date') else str(c) for c in df.columns]
            return df.to_dict()

        return {
            "info": {
                "name":               info.get("longName", ""),
                "sector":             info.get("sector", ""),
                "industry":           info.get("industryKey", ""),
                "market_cap_cr":      round((info.get("marketCap", 0) or 0) / 1e7, 0),
                "current_price":      info.get("currentPrice", info.get("regularMarketPrice", 0)),
                "52w_high":           info.get("fiftyTwoWeekHigh", 0),
                "52w_low":            info.get("fiftyTwoWeekLow", 0),
                "book_value":         info.get("bookValue", 0),
                "price_to_book":      info.get("priceToBook", 0),
                "trailing_pe":        info.get("trailingPE", 0),
                "forward_pe":         info.get("forwardPE", 0),
                "dividend_yield_pct": round(((info.get("dividendYield") or 0)) * 100, 2),
                "roe_pct":            round(((info.get("returnOnEquity") or 0)) * 100, 2),
                "roa_pct":            round(((info.get("returnOnAssets") or 0)) * 100, 2),
                "shares_outstanding": info.get("sharesOutstanding", 0),
                "beta":               info.get("beta", 1.0),
                "description":        (info.get("longBusinessSummary") or "")[:500],
            },
            "income_statement":  df_to_dict(stock.financials),
            "balance_sheet":     df_to_dict(stock.balance_sheet),
            "cashflow":          df_to_dict(stock.cashflow),
            "quarterly_income":  df_to_dict(stock.quarterly_financials),
        }
    except Exception as e:
        print(f"  ⚠️  yfinance error: {e}")
        return {}


def fetch_screener_data(ticker: str) -> str:
    try:
        import requests
        from bs4 import BeautifulSoup
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        }
        for url in [
            f"https://www.screener.in/company/{ticker}/consolidated/",
            f"https://www.screener.in/company/{ticker}/",
        ]:
            r = requests.get(url, headers=headers, timeout=15)
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, "html.parser")
                rows = []
                for table in soup.find_all("section", class_="card"):
                    h = table.find("h2")
                    if h:
                        rows.append(f"\n=== {h.get_text(strip=True)} ===")
                    for row in table.find_all("tr"):
                        cells = row.find_all(["th", "td"])
                        if cells:
                            rows.append(" | ".join(c.get_text(strip=True) for c in cells))
                if rows:
                    return "\n".join(rows)
        return "screener.in: no data found"
    except Exception as e:
        return f"screener.in error: {e}"


# ══════════════════════════════════════════════════════════════════════════════
# LLM UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def _retry(fn, max_retries=5):
    for attempt in range(max_retries):
        try:
            return fn()
        except anthropic.RateLimitError:
            if attempt == max_retries - 1:
                raise
            wait = 60 * (attempt + 1)
            print(f"⏳ Rate limited, waiting {wait}s...")
            time.sleep(wait)


def extract_text(response) -> str:
    return "".join(
        b.text for b in response.content
        if hasattr(b, "text") and b.text
    )


def extract_json(text: str) -> dict:
    """Robustly extract JSON from LLM response."""
    # Strip thinking tags
    text = re.sub(r'<thinking>.*?</thinking>', '', text, flags=re.DOTALL).strip()

    # Direct parse
    if text.startswith('{'):
        try:
            return json.loads(text)
        except Exception:
            pass

    # JSON code block
    for pat in [r'```json\s*(.*?)\s*```', r'```\s*(\{.*?\})\s*```']:
        m = re.search(pat, text, re.DOTALL)
        if m:
            try:
                return json.loads(m.group(1))
            except Exception:
                pass

    # Largest { ... } block
    start = text.find('{')
    if start != -1:
        depth = 0
        for i, ch in enumerate(text[start:], start):
            if ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[start:i + 1])
                    except Exception:
                        break

    raise ValueError(f"No valid JSON found. Preview: {text[:300]}")


def run_agent(prompt: str, tools=None, max_tokens: int = 16000) -> str:
    """Agentic loop — handles web_search tool calls until end_turn."""
    messages = [{"role": "user", "content": prompt}]

    for _ in range(30):  # safety cap
        kwargs = dict(model=SONNET, max_tokens=max_tokens, messages=messages)
        if tools:
            kwargs["tools"] = tools

        response = _retry(lambda: client.messages.create(**kwargs))

        if response.stop_reason == "end_turn":
            return extract_text(response)

        if response.stop_reason == "tool_use":
            messages.append({"role": "assistant", "content": response.content})
            tool_results = [
                {"type": "tool_result", "tool_use_id": b.id, "content": ""}
                for b in response.content
                if hasattr(b, "type") and b.type == "tool_use"
            ]
            if tool_results:
                messages.append({"role": "user", "content": tool_results})
        else:
            return extract_text(response)

    return extract_text(response)


# ══════════════════════════════════════════════════════════════════════════════
# CALL 1 — FULL ANALYSIS → JSON
# ══════════════════════════════════════════════════════════════════════════════

JSON_SCHEMA = """
{
  "company": {
    "name": str, "ticker": str, "exchange": str, "sector": str,
    "sector_type": "banking OR non_banking",
    "cmp": number, "market_cap_cr": number, "shares_outstanding_cr": number,
    "book_value_per_share": number, "52w_high": number, "52w_low": number,
    "beta": number, "description": str
  },
  "data_quality": {
    "overall": "High|Medium|Low", "sources_used": [str],
    "gaps": [str], "notes": str
  },
  "historical": {
    "years": ["FY20","FY21","FY22","FY23","FY24","FY25"],
    "income_statement": {
      "nii": [6 values],           "other_income": [6],
      "total_income": [6],         "operating_expenses": [6],
      "ppop": [6],                 "provisions": [6],
      "revenue": [6],              "cogs": [6],
      "gross_profit": [6],         "ebitda": [6],
      "depreciation": [6],         "ebit": [6],
      "interest_expense": [6],     "pbt": [6],
      "tax": [6],                  "pat": [6],
      "eps_basic": [6],            "dps": [6]
    },
    "balance_sheet": {
      "total_assets": [6],         "gross_advances": [6],
      "net_advances": [6],         "investments": [6],
      "cash": [6],                 "total_deposits": [6],
      "casa_deposits": [6],        "borrowings": [6],
      "other_liabilities": [6],    "total_liabilities": [6],
      "fixed_assets_net": [6],     "current_assets": [6],
      "inventory": [6],            "receivables": [6],
      "total_debt": [6],           "current_liabilities": [6],
      "share_capital": [6],        "reserves": [6],
      "total_equity": [6],
      "gnpa_cr": [6],              "nnpa_cr": [6]
    },
    "cash_flow": {
      "cfo": [6], "cfi": [6], "cff": [6],
      "net_change_in_cash": [6], "capex": [6], "fcf": [6]
    },
    "key_ratios": {
      "roe_pct": [6],              "roa_pct": [6],
      "pat_margin_pct": [6],       "nim_pct": [6],
      "gnpa_pct": [6],             "nnpa_pct": [6],
      "pcr_pct": [6],              "cost_to_income_pct": [6],
      "car_pct": [6],              "tier1_pct": [6],
      "casa_ratio_pct": [6],       "credit_cost_pct": [6],
      "ebitda_margin_pct": [6],    "gross_margin_pct": [6],
      "roce_pct": [6],             "debt_to_equity": [6],
      "interest_coverage": [6]
    }
  },
  "assumptions": {
    "projection_years": ["FY26E","FY27E","FY28E","FY29E","FY30E"],
    "macro": {
      "rbi_repo_rate_pct": {"base": n, "bull": n, "bear": n},
      "india_gdp_growth_pct": {"base": n, "bull": n, "bear": n},
      "gsec_10yr_yield_pct": number
    },
    "revenue_drivers": {
      "loan_growth_pct":    {"FY26E":{"base":n,"bull":n,"bear":n}, ...5 years},
      "nim_bps":            {"FY26E":{"base":n,"bull":n,"bear":n}, ...},
      "fee_income_growth":  {"FY26E":{"base":n,"bull":n,"bear":n}, ...},
      "revenue_growth_pct": {"FY26E":{"base":n,"bull":n,"bear":n}, ...},
      "ebitda_margin_pct":  {"FY26E":{"base":n,"bull":n,"bear":n}, ...}
    },
    "cost_drivers": {
      "cost_to_income_pct": {"FY26E":{"base":n,"bull":n,"bear":n}, ...},
      "credit_cost_pct":    {"FY26E":{"base":n,"bull":n,"bear":n}, ...},
      "gnpa_pct":           {"FY26E":{"base":n,"bull":n,"bear":n}, ...},
      "opex_growth_pct":    {"FY26E":{"base":n,"bull":n,"bear":n}, ...}
    },
    "capital_structure": {
      "tax_rate_pct": number, "dividend_payout_pct": number,
      "shares_outstanding_cr": number
    },
    "rationale": {
      "key_assumptions": str, "bull_drivers": str, "bear_risks": str
    }
  },
  "projections": {
    "years": ["FY26E","FY27E","FY28E","FY29E","FY30E"],
    "income_statement": {
      "base": {same IS keys as historical, each with 5 values},
      "bull": {same},
      "bear": {same}
    },
    "balance_sheet": {
      "base": {same BS keys as historical, each with 5 values},
      "bull": {same},
      "bear": {same}
    },
    "cash_flow": {
      "base": {same CF keys, 5 values each},
      "bull": {same},
      "bear": {same}
    },
    "key_ratios": {
      "base": {same ratio keys, 5 values each},
      "bull": {same},
      "bear": {same}
    },
    "cagr_summary": {
      "pat_fy22_fy25_pct": n, "pat_fy25_fy30_base_pct": n,
      "eps_fy25_fy30_base_pct": n, "bvps_fy25_fy30_base_pct": n
    }
  },
  "wacc": {
    "risk_free_rate_pct": n, "equity_risk_premium_pct": n,
    "beta": n, "cost_of_equity_pct": n,
    "cost_of_debt_pre_tax_pct": n, "tax_rate_pct": n,
    "cost_of_debt_post_tax_pct": n, "debt_weight_pct": n,
    "equity_weight_pct": n, "wacc_pct": n,
    "terminal_growth_rate_pct": n, "wacc_rationale": str
  },
  "valuation": {
    "dcf": {
      "fcfe_base_cr": [5 values FY26E-FY30E],
      "terminal_value_cr": n, "pv_fcfe_cr": n,
      "pv_terminal_value_cr": n, "total_equity_value_cr": n,
      "fair_value_per_share": n, "weight_pct": n
    },
    "pe_comps": {
      "peers": [{"name": str, "pe_ttm": n}],
      "sector_median_pe": n, "target_pe": n,
      "base_year": "FY27E", "base_year_eps": n,
      "fair_value_per_share": n, "weight_pct": n
    },
    "pbv_comps": {
      "peers": [{"name": str, "pbv": n}],
      "sector_median_pbv": n, "target_pbv": n,
      "base_year": "FY27E", "base_year_bvps": n,
      "fair_value_per_share": n, "weight_pct": n
    },
    "ggm": {
      "cost_of_equity_pct": n, "sustainable_roe_pct": n,
      "terminal_growth_pct": n, "implied_pbv": n,
      "base_year_bvps": n, "fair_value_per_share": n, "weight_pct": n
    },
    "blended": {
      "fair_value_per_share": n, "cmp": n, "upside_pct": n,
      "rating": "STRONG BUY|BUY|ACCUMULATE|HOLD|UNDERPERFORM|SELL",
      "target_price_12m": n
    },
    "scenario_targets": {"bull": n, "base": n, "bear": n}
  },
  "sensitivity": {
    "dcf_grid": {
      "row_label": "WACC (%)", "col_label": "Terminal Growth Rate (%)",
      "row_values": [9.0,9.5,10.0,10.5,11.0],
      "col_values": [4.0,4.5,5.0,5.5,6.0],
      "grid": [[5x5 fair value numbers]]
    },
    "pe_grid": {
      "row_label": "EPS CAGR (%)", "col_label": "Target P/E",
      "row_values": [12,15,18,21,24], "col_values": [14,16,18,20,22],
      "grid": [[5x5 fair value numbers]]
    },
    "pbv_grid": {
      "row_label": "ROE (%)", "col_label": "Target P/BV",
      "row_values": [14,15,16,17,18], "col_values": [2.5,2.75,3.0,3.25,3.5],
      "grid": [[5x5 fair value numbers]]
    }
  }
}
"""


def run_analysis(company_name: str, ticker: str, exchange: str, sector: str,
                 yf_data: dict, screener_text: str) -> dict:
    """Call 1 — full research, historical, projections, valuation → JSON."""

    banking_kw = ["bank", "nbfc", "financial", "insurance", "lending", "microfinance", "housing finance"]
    is_banking = any(k in sector.lower() for k in banking_kw)
    sector_type = "banking" if is_banking else "non_banking"

    yf_json  = json.dumps(yf_data, indent=2, default=str)[:6000]
    sc_text  = screener_text[:6000]

    prompt = f"""You are a senior equity research analyst at a top-tier investment bank.
Build a complete institutional-quality financial model for {company_name} ({ticker}, {exchange}).
Sector: {sector}  |  Sector type: {sector_type}

━━ PRE-FETCHED DATA ━━

SOURCE 1 — Yahoo Finance:
{yf_json}

SOURCE 2 — Screener.in (10-yr audited):
{sc_text}

━━ EXECUTE THESE STEPS IN ORDER ━━

STEP 1 — LIVE WEB SEARCHES (do all of these):
  • "{ticker} site:screener.in financial statements 10 year"
  • "{company_name} quarterly results FY25 earnings PAT NII"
  • "{company_name} management guidance FY26 outlook annual report"
  • "{sector} sector NSE peer companies P/E P/BV multiples 2025"
  • "India G-Sec 10 year yield current 2025"
  • "{company_name} analyst target price Bloomberg consensus 2025"

STEP 2 — HISTORICAL DATA (FY20–FY25):
  • Screener.in is primary (audited). yfinance is secondary.
  • For banking: NII, Other Income, PPOP, Provisions, PBT, Tax, PAT, EPS, DPS
    BS: Total Assets, Advances, Investments, Deposits, CASA, Equity, GNPA, NNPA
    Ratios: NIM, GNPA%, NNPA%, PCR, Cost/Income, CAR, Tier1, CASA%, Credit Cost
  • For non-banking: Revenue, EBITDA, EBIT, PBT, PAT, EPS, DPS
    BS: Fixed Assets, Inventory, Receivables, Debt, Equity, Current Assets/Liabilities
    Ratios: EBITDA margin, PAT margin, ROCE, D/E, Interest coverage
  • Use null for genuinely missing values — NEVER estimate historical numbers
  • VERIFY: Total Assets = Total Liabilities + Total Equity for every year

STEP 3 — ASSUMPTIONS (FY26E–FY30E, Bull/Base/Bear):
  • Every assumption MUST cite evidence (mgmt guidance, peer data, RBI/macro data)
  • Banking: Loan growth vs RBI system credit growth (typically 12-16%)
    NIM trajectory (repo rate cycle impact, MCLR/EBLR mix)
    Credit cost glide path (post-NPA cycle normalisation)
    GNPA% target (management guidance), Cost/Income efficiency
  • Non-banking: Revenue growth drivers (volume + price), margin expansion path
    Working capital trends, capex cycle, debt reduction
  • Bull: company executes perfectly + favorable macro
  • Bear: execution miss + adverse macro (rate hike, NPA cycle, demand slowdown)

STEP 4 — PROJECT FINANCIALS (FY26E–FY30E, all 3 scenarios):
  • Each line MUST mathematically follow from assumptions
  • BS MUST balance: Assets = Liabilities + Equity — verify before outputting
  • CF uses indirect method: CFO = PAT + D&A + Provisions ± WC changes
  • EPS = PAT / shares outstanding (account for dilution if any)
  • Show realistic YoY variation — not flat growth rates every year

STEP 5 — WACC (use live G-Sec yield from Step 1):
  • Rf = India G-Sec 10yr (exact current value from web search)
  • ERP = India ERP from Damodaran (typically 5.0–6.5%)
  • Beta = from yfinance or 3yr regression vs Nifty 50
  • Kd = weighted avg borrowing cost (from BS interest expense / total debt)
  • Terminal growth = India long-term nominal GDP (7–8%)

STEP 6 — VALUATION (4 methods):
  • DCF: FCFE = PAT + D&A - Capex ± WC changes - Net debt repayment
    Terminal value = FCFE_terminal / (Ke - g)  [Gordon Growth]
    Discount each year's FCFE at Ke, not WACC (equity model)
  • P/E Comps: 3–5 sector peers from web search → sector median → justified target
  • P/BV Comps: same methodology
  • GGM (esp banking): Implied P/BV = (ROE - g) / (Ke - g); fair value = P/BV × BVPS
  • Weights: DCF 40%, P/E 25%, P/BV 25%, GGM 10%
  • Rating: STRONG BUY (>20% upside) | BUY (10-20%) | ACCUMULATE (5-10%)
            HOLD (-5 to 5%) | UNDERPERFORM (-10 to -5%) | SELL (<-10%)

STEP 7 — SENSITIVITY (3 grids, 5×5 each):
  • DCF grid: WACC axis [Ke-200bps to Ke+200bps in 50bps steps] vs TGR [4-6% in 50bps]
  • P/E grid: EPS CAGR [12,15,18,21,24%] vs Target P/E [14,16,18,20,22x]
  • P/BV grid: ROE [14,15,16,17,18%] vs Target P/BV [2.5,2.75,3.0,3.25,3.5x]

━━ ABSOLUTE RULES ━━
1. All monetary values in ₹ Crores. EPS, BVPS, CMP, Target Price in ₹ per share.
2. Percentages as plain numbers: write 15.5 not 0.155
3. null for missing data. NEVER write 0 when data is genuinely unavailable.
4. BS MUST balance every year — if it doesn't, fix it before outputting.
5. Projections must be internally consistent with assumptions.
6. Use sector_type = "{sector_type}"

━━ OUTPUT ━━
Return ONLY a valid JSON object. Start with {{ end with }}. No prose, no markdown fences.
Match this schema exactly:

{JSON_SCHEMA}
"""

    print("🤖 Call 1 | Research + Analysis + Projections + Valuation → JSON ...")
    raw = run_agent(prompt, tools=WEB_SEARCH_TOOL, max_tokens=16000)

    try:
        return extract_json(raw)
    except ValueError as e:
        print(f"  ⚠️  JSON parse failed: {e}")
        print("  🔄 Requesting JSON repair...")
        repair_prompt = f"""The text below contains financial model data but is not valid JSON.
Extract all data and return ONLY a valid JSON object starting with {{ and ending with }}.

TEXT:
{raw[:12000]}"""
        repaired = _retry(lambda: client.messages.create(
            model=SONNET, max_tokens=16000,
            messages=[{"role": "user", "content": repair_prompt}]
        ))
        return extract_json(extract_text(repaired))


# ══════════════════════════════════════════════════════════════════════════════
# CALL 2 — INVESTMENT THESIS → JSON
# ══════════════════════════════════════════════════════════════════════════════

def run_thesis(company_name: str, ticker: str, model_json: dict) -> dict:
    """Call 2 — generate SAARTHI-scored investment thesis from valuation JSON."""

    val   = model_json.get("valuation", {})
    blend = val.get("blended", {})
    proj  = model_json.get("projections", {})
    asmp  = model_json.get("assumptions", {})

    prompt = f"""You are a senior equity research analyst writing the investment thesis
for the cover page of a professional financial model.

VALUATION OUTPUT:
  Company:       {company_name} ({ticker})
  CMP:           ₹{blend.get('cmp', 'NA')}
  Target Price:  ₹{blend.get('target_price_12m', 'NA')}
  Upside:        {blend.get('upside_pct', 'NA')}%
  Rating:        {blend.get('rating', 'NA')}
  Bull target:   ₹{val.get('scenario_targets', {}).get('bull', 'NA')}
  Bear target:   ₹{val.get('scenario_targets', {}).get('bear', 'NA')}

PROJECTION CAGRS:
{json.dumps(proj.get('cagr_summary', {}), indent=2)}

ASSUMPTION RATIONALE:
{json.dumps(asmp.get('rationale', {}), indent=2)}

SAARTHI FRAMEWORK (score each dimension 0 to max points):
  S — Sector & Business Quality    (0–15 pts): industry tailwinds, moat, competitive position
  A — Accounting Quality           (0–15 pts): earnings quality, cash conversion, governance
  A — Asset Quality                (0–15 pts): for banking GNPA/NNPA/PCR; else asset efficiency
  R — Revenue Visibility           (0–15 pts): order book, pricing power, recurring revenues
  T — Track Record                 (0–10 pts): mgmt execution history, ROE/ROA consistency
  H — Health of Balance Sheet      (0–15 pts): leverage, liquidity, capital adequacy
  I — Intrinsic Valuation          (0–15 pts): upside from blended fair value vs CMP
  TOTAL = sum of above (max 100)
  Rating: STRONG BUY ≥80 | BUY 65–79 | ACCUMULATE 55–64 | HOLD 45–54 |
          UNDERPERFORM 35–44 | SELL <35

Return ONLY this JSON object (no prose, no markdown):
{{
  "rating": "{blend.get('rating', 'HOLD')}",
  "target_price": {blend.get('target_price_12m', 0)},
  "cmp": {blend.get('cmp', 0)},
  "upside_pct": {blend.get('upside_pct', 0)},
  "saarthi_scores": {{
    "S_sector_quality": <0-15>,
    "A_accounting_quality": <0-15>,
    "A_asset_quality": <0-15>,
    "R_revenue_visibility": <0-15>,
    "T_track_record": <0-10>,
    "H_balance_sheet_health": <0-15>,
    "I_intrinsic_valuation": <0-15>
  }},
  "saarthi_total": <integer 0-100>,
  "saarthi_rating": "<STRONG BUY|BUY|ACCUMULATE|HOLD|UNDERPERFORM|SELL>",
  "investment_thesis": "<150-word thesis: core investment case, key value driver, why now>",
  "bull_case": "<80-word bull case: what executes better than expected>",
  "bear_case": "<80-word bear case: key risks and downside scenario>",
  "key_catalysts": ["<catalyst 1>", "<catalyst 2>", "<catalyst 3>", "<catalyst 4>"],
  "key_risks": ["<risk 1>", "<risk 2>", "<risk 3>", "<risk 4>"]
}}
"""

    print("📝 Call 2 | Investment thesis (SAARTHI) ...")
    response = _retry(lambda: client.messages.create(
        model=SONNET, max_tokens=1500,
        messages=[{"role": "user", "content": prompt}]
    ))
    raw = extract_text(response)
    try:
        return extract_json(raw)
    except Exception:
        return {
            "rating": blend.get("rating", "HOLD"),
            "target_price": blend.get("target_price_12m", 0),
            "cmp": blend.get("cmp", 0),
            "upside_pct": blend.get("upside_pct", 0),
            "saarthi_scores": {},
            "saarthi_total": 60,
            "saarthi_rating": blend.get("rating", "HOLD"),
            "investment_thesis": raw[:500],
            "bull_case": "See assumptions sheet.",
            "bear_case": "See assumptions sheet.",
            "key_catalysts": ["Earnings growth", "Sector tailwinds", "Management execution", "Re-rating"],
            "key_risks": ["Macro headwinds", "Competition", "Regulatory risk", "Execution miss"],
        }


# ══════════════════════════════════════════════════════════════════════════════
# DETERMINISTIC EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(model: dict, thesis: dict, out_path: str, company_name: str, ticker: str):
    """Build a professional 10-sheet Excel model from JSON. No LLM required."""

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    # ── Colour palette ──────────────────────────────────────────────────────
    C = {
        "dark_blue":  "1F3864", "med_blue":   "2E75B6",
        "light_blue": "BDD7EE", "very_light": "DEEAF1",
        "white":      "FFFFFF", "gray":       "F2F2F2",
        "dark_gray":  "D9D9D9", "black":      "000000",
        "green":      "E2EFDA", "dark_green": "375623",
        "red_light":  "FFE0CC", "dark_red":   "C00000",
        "yellow":     "FFF2CC", "gold":       "BF8F00",
        "teal":       "1F6B75",
    }

    # ── Style helpers ───────────────────────────────────────────────────────
    def _fill(k): return PatternFill("solid", fgColor=C.get(k, k))
    def _font(bold=False, color="black", size=10, italic=False):
        return Font(name="Calibri", size=size, bold=bold,
                    color=C.get(color, color), italic=italic)
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def _side(s="thin"): return Side(style=s)
    def _border(s="thin"):
        sd = _side(s)
        return Border(left=sd, right=sd, top=sd, bottom=sd)

    def w(ws, row, col, value, bold=False, bg=None, fg="black", size=10,
          h="left", wrap=False, italic=False, fmt=None, border=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font  = _font(bold=bold, color=fg, size=size, italic=italic)
        cell.alignment = _align(h=h, wrap=wrap)
        if bg:
            cell.fill = _fill(bg)
        if border:
            cell.border = _border()
        if fmt:
            cell.number_format = fmt
        return cell

    def hdr(ws, row, cols_labels, c0=1, bg="dark_blue", fg="white", size=10):
        for i, lbl in enumerate(cols_labels):
            w(ws, row, c0 + i, lbl, bold=True, bg=bg, fg=fg, size=size, h="center")

    def merge_hdr(ws, row, c1, c2, text, bg="med_blue", fg="white"):
        cell = ws.cell(row=row, column=c1, value=text)
        cell.font  = _font(bold=True, color=fg, size=10)
        cell.fill  = _fill(bg)
        cell.alignment = _align(h="left")
        if c2 > c1:
            ws.merge_cells(start_row=row, start_column=c1,
                           end_row=row, end_column=c2)

    def col_w(ws, pairs):
        for col, width in pairs:
            ws.column_dimensions[get_column_letter(col)].width = width

    def v(val, default=None):
        """Safe value — return default if val is None."""
        return default if val is None else val

    def pct_str(val):
        if val is None: return "NA"
        try: return f"{float(val):.1f}%"
        except: return str(val)

    def num_str(val, dec=1):
        if val is None: return "NA"
        try: return f"{float(val):,.{dec}f}"
        except: return str(val)

    # ── Data shortcuts ──────────────────────────────────────────────────────
    hist       = model.get("historical", {})
    hist_yrs   = hist.get("years", ["FY20","FY21","FY22","FY23","FY24","FY25"])
    proj       = model.get("projections", {})
    proj_yrs   = proj.get("years", ["FY26E","FY27E","FY28E","FY29E","FY30E"])
    all_yrs    = hist_yrs + proj_yrs
    nh, np_    = len(hist_yrs), len(proj_yrs)

    hist_is    = hist.get("income_statement", {})
    hist_bs    = hist.get("balance_sheet", {})
    hist_cf    = hist.get("cash_flow", {})
    hist_kr    = hist.get("key_ratios", {})

    proj_is    = proj.get("income_statement", {})
    proj_bs    = proj.get("balance_sheet", {})
    proj_cf    = proj.get("cash_flow", {})
    proj_kr    = proj.get("key_ratios", {})

    wacc_d     = model.get("wacc", {})
    val_d      = model.get("valuation", {})
    sens_d     = model.get("sensitivity", {})
    asmp       = model.get("assumptions", {})
    co         = model.get("company", {})

    is_banking = co.get("sector_type", "banking") == "banking"
    blend      = val_d.get("blended", {})
    rating     = thesis.get("rating", blend.get("rating", "HOLD"))

    RATING_COLOR = {
        "STRONG BUY": "dark_green", "BUY": "dark_green",
        "ACCUMULATE": "gold", "HOLD": "gold",
        "UNDERPERFORM": "dark_red", "SELL": "dark_red",
    }
    rc = RATING_COLOR.get(rating, "black")

    def series(hist_d, proj_d_scenario, key, n_h=nh, n_p=np_):
        """Combine historical + projected series for a key."""
        h = (hist_d.get(key) or []) + [None] * n_h
        p = (proj_d_scenario.get(key) or []) + [None] * n_p
        return h[:n_h] + p[:n_p]

    FMT_CR  = '#,##0.0'
    FMT_CR0 = '#,##0'
    FMT_PCT = '0.00'
    FMT_EPS = '#,##0.00'
    FMT_RS  = '₹#,##0'
    FMT_X   = '0.0x'

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1: COVER
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    col_w(ws, [(1,3),(2,30),(3,20),(4,20),(5,20),(6,20)])

    # Title band
    for col in range(1, 7):
        ws.cell(row=1, column=col).fill = _fill("dark_blue")
        ws.cell(row=2, column=col).fill = _fill("dark_blue")
        ws.cell(row=3, column=col).fill = _fill("dark_blue")

    w(ws, 2, 2, f"TIKONA CAPITAL  |  EQUITY RESEARCH",
      bold=True, fg="white", size=14, h="left")
    w(ws, 3, 2, f"Financial Model  —  {company_name} ({ticker})",
      fg="light_blue", size=11, italic=True)

    r = 5
    w(ws, r, 2, "INVESTMENT SUMMARY", bold=True, fg="dark_blue", size=12)
    r += 1

    def kv(row, label, value, val_bold=False, val_color="black", val_fmt=None):
        w(ws, row, 2, label, fg="dark_blue", bold=True, size=10)
        cell = w(ws, row, 3, value, bold=val_bold, fg=val_color, size=10, h="right")
        if val_fmt:
            cell.number_format = val_fmt

    kv(r,   "Company",        co.get("name", company_name));               r+=1
    kv(r,   "NSE Symbol",     ticker);                                      r+=1
    kv(r,   "Sector",         co.get("sector", ""));                        r+=1
    kv(r,   "CMP (₹)",        v(blend.get("cmp"), 0),   val_fmt=FMT_RS);   r+=1
    kv(r,   "Target Price (₹)",v(blend.get("target_price_12m"),0), val_bold=True, val_color=rc, val_fmt=FMT_RS); r+=1
    kv(r,   "Upside / (Downside)", f"{v(blend.get('upside_pct'),0):.1f}%", val_bold=True, val_color=rc); r+=1
    kv(r,   "Rating",         rating, val_bold=True, val_color=rc);          r+=1
    kv(r,   "Market Cap (₹ Cr)", v(co.get("market_cap_cr"),0), val_fmt=FMT_CR0); r+=1
    kv(r,   "52-Week High (₹)", v(co.get("52w_high"),0), val_fmt=FMT_RS);   r+=1
    kv(r,   "52-Week Low (₹)", v(co.get("52w_low"),0), val_fmt=FMT_RS);     r+=1
    kv(r,   "Beta",            v(co.get("beta"),1.0));                       r+=2

    # Scenario targets
    sc_targets = val_d.get("scenario_targets", {})
    w(ws, r, 2, "SCENARIO TARGET PRICES (12M)", bold=True, fg="dark_blue", size=11); r+=1
    for sc, col_idx, color in [("Bull", 2, "dark_green"), ("Base", 3, "black"), ("Bear", 4, "dark_red")]:
        w(ws, r, col_idx, f"{sc} Case: ₹{v(sc_targets.get(sc.lower()), 0):,.0f}",
          bold=(sc=="Base"), fg=color)
    r += 2

    # SAARTHI scorecard
    scores = thesis.get("saarthi_scores", {})
    total  = thesis.get("saarthi_total", 0)
    s_dims = [
        ("S", "Sector & Business Quality", "S_sector_quality", 15),
        ("A", "Accounting Quality",         "A_accounting_quality", 15),
        ("A", "Asset Quality",              "A_asset_quality", 15),
        ("R", "Revenue Visibility",         "R_revenue_visibility", 15),
        ("T", "Track Record",               "T_track_record", 10),
        ("H", "Balance Sheet Health",       "H_balance_sheet_health", 15),
        ("I", "Intrinsic Valuation",        "I_intrinsic_valuation", 15),
    ]
    w(ws, r, 2, "SAARTHI SCORECARD", bold=True, fg="dark_blue", size=11); r+=1
    hdr(ws, r, ["Dimension", "Description", "Score", "Max"], c0=2, bg="med_blue"); r+=1
    for letter, desc, key, mx in s_dims:
        sc_val = scores.get(key, "—")
        bg = "green" if (isinstance(sc_val, (int,float)) and sc_val >= mx*0.7) else "gray"
        for ci, val in enumerate([letter, desc, sc_val, mx], 2):
            w(ws, r, ci, val, bg=bg if ci==4 else None, h="center" if ci in [3,4] else "left")
        r += 1
    w(ws, r, 2, "TOTAL", bold=True)
    w(ws, r, 3, total, bold=True, fg=rc, h="center")
    w(ws, r, 4, 100,   bold=True, h="center")
    r += 2

    # Thesis
    w(ws, r, 2, "INVESTMENT THESIS", bold=True, fg="dark_blue", size=11); r+=1
    cell = ws.cell(row=r, column=2, value=thesis.get("investment_thesis", ""))
    cell.font = _font(size=10, italic=True)
    cell.alignment = _align(wrap=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r+4, end_column=6)
    ws.row_dimensions[r].height = 80
    r += 6

    w(ws, r, 2, "KEY CATALYSTS", bold=True, fg="dark_green"); r+=1
    for cat in thesis.get("key_catalysts", []):
        w(ws, r, 2, f"▶  {cat}"); r+=1
    r += 1
    w(ws, r, 2, "KEY RISKS", bold=True, fg="dark_red"); r+=1
    for risk in thesis.get("key_risks", []):
        w(ws, r, 2, f"▶  {risk}"); r+=1

    ws.freeze_panes = "B4"

    # ════════════════════════════════════════════════════════════════════════
    # HELPER: write a financial table (IS / BS / CF / Ratios)
    # ════════════════════════════════════════════════════════════════════════
    def write_fin_table(ws, start_row, title, rows_def, base, bull=None, bear=None,
                        show_hist=True, show_base=True, show_bull=True, show_bear=True):
        """
        rows_def: list of (label, key, fmt, is_section_header, indent)
        base: dict with key → list of values
        Returns next available row.
        """
        COL_LABEL = 1
        col_data_start = 2

        # Column headers
        labels_row = []
        data_cols  = []  # (col_idx, data_source_dict, value_idx)

        col = col_data_start
        if show_hist:
            for i, yr in enumerate(hist_yrs):
                labels_row.append(yr)
                data_cols.append(("hist", i))
                col += 1
        if show_base:
            for i, yr in enumerate(proj_yrs):
                labels_row.append(yr + "E" if "E" not in yr else yr)
                data_cols.append(("base", i))
                col += 1
        if show_bull and bull:
            for i, yr in enumerate(proj_yrs):
                labels_row.append(yr + " Bull")
                data_cols.append(("bull", i))
                col += 1
        if show_bear and bear:
            for i, yr in enumerate(proj_yrs):
                labels_row.append(yr + " Bear")
                data_cols.append(("bear", i))
                col += 1

        r = start_row
        merge_hdr(ws, r, COL_LABEL, COL_LABEL + len(labels_row), title, bg="dark_blue")
        r += 1

        hdr_labels = ["(₹ Crore unless noted)"] + labels_row
        hdr(ws, r, hdr_labels, c0=COL_LABEL, bg="med_blue"); r += 1

        # Separator lines between hist and proj
        hist_sep_col = col_data_start + nh if show_hist else None

        for label, key, fmt, is_sec, indent in rows_def:
            if is_sec:
                merge_hdr(ws, r, COL_LABEL, COL_LABEL + len(labels_row), f"  {label}",
                          bg="very_light", fg="dark_blue"); r += 1
                continue

            prefix = "  " * indent
            w(ws, r, COL_LABEL, prefix + label, bold=(indent == 0))

            for ci, (src, idx) in enumerate(data_cols):
                col_idx = col_data_start + ci
                if src == "hist":
                    data = base  # hist data already combined or passed as hist
                    val = (base.get(key) or [None]*nh)[idx] if src == "hist" else None
                    # Re-map: base here is the historical dict for hist source
                elif src == "base":
                    val_list = (base.get(key) or [None]*np_)
                    val = val_list[idx] if idx < len(val_list) else None
                elif src == "bull":
                    val_list = (bull.get(key) or [None]*np_) if bull else [None]*np_
                    val = val_list[idx] if idx < len(val_list) else None
                elif src == "bear":
                    val_list = (bear.get(key) or [None]*np_) if bear else [None]*np_
                    val = val_list[idx] if idx < len(val_list) else None
                else:
                    val = None

                bg_color = None
                if ci > 0 and hist_sep_col and col_idx == hist_sep_col:
                    # projected area
                    bg_color = "very_light" if r % 2 == 0 else None
                elif r % 2 == 0:
                    bg_color = "gray"

                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.font = _font()
                cell.alignment = _align(h="right")
                if bg_color:
                    cell.fill = _fill(bg_color)
                if fmt:
                    cell.number_format = fmt

            r += 1

        return r + 1

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2: INCOME STATEMENT
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Income Statement")
    ws.freeze_panes = "B3"
    col_w(ws, [(1,32)] + [(i, 12) for i in range(2, 18)])

    r = 1
    # Title
    w(ws, r, 1, f"INCOME STATEMENT  —  {company_name} ({ticker})",
      bold=True, bg="dark_blue", fg="white", size=12); r += 1

    years_hdr = ["(₹ Crore)"] + all_yrs
    hdr(ws, r, years_hdr, c0=1, bg="med_blue"); r += 1

    def write_is_row(label, key, fmt=FMT_CR, bold=False, bg_label=None, separator=False):
        nonlocal r
        if separator:
            for ci in range(1, len(all_yrs)+2):
                ws.cell(row=r, column=ci).fill = _fill("dark_gray")
            r += 1; return

        w(ws, r, 1, label, bold=bold, bg=bg_label)
        h_vals = hist_is.get(key) or [None]*nh
        b_vals = proj_is.get("base", {}).get(key) or [None]*np_
        for ci, val in enumerate(h_vals[:nh] + b_vals[:np_], 2):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = _font(bold=bold)
            cell.alignment = _align(h="right")
            cell.number_format = fmt
            if r % 2 == 0:
                cell.fill = _fill("gray")
        r += 1

    if is_banking:
        write_is_row("Net Interest Income (NII)",  "nii",              FMT_CR, bold=True)
        write_is_row("  Other Income",              "other_income",     FMT_CR)
        write_is_row("Total Income",                "total_income",     FMT_CR, bold=True, bg_label="very_light")
        write_is_row("  Operating Expenses",        "operating_expenses",FMT_CR)
        write_is_row("Pre-Provision Op. Profit",    "ppop",             FMT_CR, bold=True, bg_label="very_light")
        write_is_row("  Provisions & Contingencies","provisions",       FMT_CR)
        write_is_row("Profit Before Tax (PBT)",     "pbt",              FMT_CR, bold=True)
        write_is_row("  Tax",                       "tax",              FMT_CR)
        write_is_row("Profit After Tax (PAT)",      "pat",              FMT_CR, bold=True, bg_label="light_blue")
        write_is_row("EPS (Basic, ₹)",              "eps_basic",        FMT_EPS, bold=True)
        write_is_row("DPS (₹)",                     "dps",              FMT_EPS)
    else:
        write_is_row("Revenue from Operations",     "revenue",          FMT_CR, bold=True)
        write_is_row("  Cost of Goods Sold",        "cogs",             FMT_CR)
        write_is_row("Gross Profit",                "gross_profit",     FMT_CR, bold=True, bg_label="very_light")
        write_is_row("  Other Operating Expenses",  "operating_expenses",FMT_CR)
        write_is_row("EBITDA",                      "ebitda",           FMT_CR, bold=True, bg_label="light_blue")
        write_is_row("  Depreciation & Amortisation","depreciation",    FMT_CR)
        write_is_row("EBIT",                        "ebit",             FMT_CR, bold=True)
        write_is_row("  Interest Expense",          "interest_expense", FMT_CR)
        write_is_row("Profit Before Tax (PBT)",     "pbt",              FMT_CR, bold=True)
        write_is_row("  Tax",                       "tax",              FMT_CR)
        write_is_row("Profit After Tax (PAT)",      "pat",              FMT_CR, bold=True, bg_label="light_blue")
        write_is_row("EPS (Basic, ₹)",              "eps_basic",        FMT_EPS, bold=True)
        write_is_row("DPS (₹)",                     "dps",              FMT_EPS)

    r += 1
    # YoY growth rows (PAT)
    w(ws, r, 1, "PAT YoY Growth (%)", bold=True, bg="very_light");
    hist_pat = hist_is.get("pat") or [None]*nh
    proj_pat = proj_is.get("base", {}).get("pat") or [None]*np_
    all_pat  = hist_pat[:nh] + proj_pat[:np_]
    for ci, (curr, prev) in enumerate(zip(all_pat[1:], all_pat[:-1]), 3):
        if curr and prev and prev != 0:
            growth = (curr - prev) / abs(prev) * 100
            cell = ws.cell(row=r, column=ci, value=round(growth, 1))
            cell.number_format = '0.0'
            cell.alignment = _align(h="right")
            cell.font = _font(color="dark_green" if growth > 0 else "dark_red")
    r += 2

    # Bull / Bear comparison
    w(ws, r, 1, "SCENARIO COMPARISON — PAT (₹ Cr)", bold=True, bg="dark_blue", fg="white"); r += 1
    hdr(ws, r, ["Scenario"] + proj_yrs, c0=1, bg="med_blue"); r += 1
    for sc_name, sc_key, fg_c in [("Base Case", "base", "black"),
                                   ("Bull Case", "bull", "dark_green"),
                                   ("Bear Case", "bear", "dark_red")]:
        vals = proj_is.get(sc_key, {}).get("pat") or [None]*np_
        w(ws, r, 1, sc_name, bold=True, fg=fg_c)
        for ci, val in enumerate(vals[:np_], 2):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.number_format = FMT_CR
            cell.alignment = _align(h="right")
            cell.font = _font(bold=True, color=fg_c)
        r += 1

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 3: BALANCE SHEET
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Balance Sheet")
    ws.freeze_panes = "B3"
    col_w(ws, [(1,32)] + [(i, 12) for i in range(2, 18)])

    r = 1
    w(ws, r, 1, f"BALANCE SHEET  —  {company_name} ({ticker})",
      bold=True, bg="dark_blue", fg="white", size=12); r += 1
    hdr(ws, r, ["(₹ Crore)"] + all_yrs, c0=1, bg="med_blue"); r += 1

    def write_bs_row(label, key, fmt=FMT_CR, bold=False, bg_label=None):
        nonlocal r
        w(ws, r, 1, label, bold=bold, bg=bg_label)
        h_vals = hist_bs.get(key) or [None]*nh
        b_vals = proj_bs.get("base", {}).get(key) or [None]*np_
        for ci, val in enumerate(h_vals[:nh] + b_vals[:np_], 2):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = _font(bold=bold)
            cell.alignment = _align(h="right")
            cell.number_format = fmt
            if r % 2 == 0:
                cell.fill = _fill("gray")
        r += 1

    def sec_hdr_bs(label):
        nonlocal r
        merge_hdr(ws, r, 1, len(all_yrs)+1, f"  {label}", bg="very_light", fg="dark_blue"); r += 1

    sec_hdr_bs("ASSETS")
    if is_banking:
        write_bs_row("Cash & Equivalents",           "cash",             FMT_CR)
        write_bs_row("Investments",                  "investments",      FMT_CR)
        write_bs_row("Gross Advances",               "gross_advances",   FMT_CR)
        write_bs_row("Net Advances",                 "net_advances",     FMT_CR, bold=True)
        write_bs_row("Other Assets",                 "other_liabilities",FMT_CR)
        write_bs_row("TOTAL ASSETS",                 "total_assets",     FMT_CR, bold=True, bg_label="light_blue")
    else:
        write_bs_row("Fixed Assets (Net)",           "fixed_assets_net", FMT_CR)
        write_bs_row("Investments",                  "investments",      FMT_CR)
        write_bs_row("Inventory",                    "inventory",        FMT_CR)
        write_bs_row("Trade Receivables",            "receivables",      FMT_CR)
        write_bs_row("Cash & Equivalents",           "cash",             FMT_CR)
        write_bs_row("Other Current Assets",         "current_assets",   FMT_CR)
        write_bs_row("TOTAL ASSETS",                 "total_assets",     FMT_CR, bold=True, bg_label="light_blue")

    r += 1
    sec_hdr_bs("LIABILITIES & EQUITY")
    if is_banking:
        write_bs_row("Total Deposits",               "total_deposits",   FMT_CR, bold=True)
        write_bs_row("  of which: CASA",             "casa_deposits",    FMT_CR)
        write_bs_row("Borrowings",                   "borrowings",       FMT_CR)
        write_bs_row("Other Liabilities",            "other_liabilities",FMT_CR)
        write_bs_row("TOTAL LIABILITIES",            "total_liabilities",FMT_CR, bold=True)
        write_bs_row("GNPA (₹ Cr)",                  "gnpa_cr",          FMT_CR)
        write_bs_row("NNPA (₹ Cr)",                  "nnpa_cr",          FMT_CR)
    else:
        write_bs_row("Total Debt",                   "total_debt",       FMT_CR, bold=True)
        write_bs_row("Trade Payables / Curr. Liab.", "current_liabilities",FMT_CR)
        write_bs_row("Other Liabilities",            "other_liabilities",FMT_CR)
        write_bs_row("TOTAL LIABILITIES",            "total_liabilities",FMT_CR, bold=True)

    r += 1
    sec_hdr_bs("SHAREHOLDERS' EQUITY")
    write_bs_row("Share Capital",                    "share_capital",    FMT_CR)
    write_bs_row("Reserves & Surplus",               "reserves",         FMT_CR)
    write_bs_row("TOTAL EQUITY",                     "total_equity",     FMT_CR, bold=True, bg_label="light_blue")

    r += 1
    # Balance check
    w(ws, r, 1, "Balance Check (Assets – Liab – Equity)", bold=True, bg="yellow");
    h_assets = hist_bs.get("total_assets") or [None]*nh
    h_liab   = hist_bs.get("total_liabilities") or [None]*nh
    h_eq     = hist_bs.get("total_equity") or [None]*nh
    p_assets = proj_bs.get("base", {}).get("total_assets") or [None]*np_
    p_liab   = proj_bs.get("base", {}).get("total_liabilities") or [None]*np_
    p_eq     = proj_bs.get("base", {}).get("total_equity") or [None]*np_
    for ci, (a, li, eq) in enumerate(
        zip(h_assets[:nh] + p_assets[:np_],
            h_liab[:nh]   + p_liab[:np_],
            h_eq[:nh]     + p_eq[:np_]), 2):
        if all(x is not None for x in [a, li, eq]):
            diff = round(a - li - eq, 1)
            cell = ws.cell(row=r, column=ci, value=diff)
            cell.number_format = FMT_CR
            cell.alignment = _align(h="right")
            cell.font = _font(color="dark_red" if abs(diff) > 1 else "dark_green", bold=True)
    r += 1

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 4: CASH FLOW
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Cash Flow")
    ws.freeze_panes = "B3"
    col_w(ws, [(1,32)] + [(i, 12) for i in range(2, 18)])

    r = 1
    w(ws, r, 1, f"CASH FLOW STATEMENT  —  {company_name} ({ticker})",
      bold=True, bg="dark_blue", fg="white", size=12); r += 1
    hdr(ws, r, ["(₹ Crore)"] + all_yrs, c0=1, bg="med_blue"); r += 1

    def write_cf_row(label, key, fmt=FMT_CR, bold=False, bg_label=None):
        nonlocal r
        w(ws, r, 1, label, bold=bold, bg=bg_label)
        h_vals = hist_cf.get(key) or [None]*nh
        b_vals = proj_cf.get("base", {}).get(key) or [None]*np_
        for ci, val in enumerate(h_vals[:nh] + b_vals[:np_], 2):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = _font(bold=bold)
            cell.alignment = _align(h="right")
            cell.number_format = fmt
            if r % 2 == 0:
                cell.fill = _fill("gray")
            if val is not None:
                try:
                    cell.font = _font(bold=bold,
                        color="dark_green" if float(val) >= 0 else "dark_red")
                except Exception:
                    pass
        r += 1

    write_cf_row("Cash Flow from Operations (CFO)", "cfo", bold=True, bg_label="light_blue")
    write_cf_row("Cash Flow from Investing (CFI)",  "cfi", bold=True)
    write_cf_row("  Capital Expenditure",           "capex")
    write_cf_row("Cash Flow from Financing (CFF)",  "cff", bold=True)
    write_cf_row("Net Change in Cash",              "net_change_in_cash", bold=True, bg_label="very_light")
    r += 1
    write_cf_row("Free Cash Flow (FCF)",            "fcf", bold=True, bg_label="green")

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 5: KEY RATIOS
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Key Ratios")
    ws.freeze_panes = "B3"
    col_w(ws, [(1,32)] + [(i, 12) for i in range(2, 18)])

    r = 1
    w(ws, r, 1, f"KEY RATIOS  —  {company_name} ({ticker})",
      bold=True, bg="dark_blue", fg="white", size=12); r += 1
    hdr(ws, r, ["Ratio"] + all_yrs, c0=1, bg="med_blue"); r += 1

    def write_kr_row(label, key, fmt=FMT_PCT, bold=False):
        nonlocal r
        w(ws, r, 1, label, bold=bold)
        h_vals = hist_kr.get(key) or [None]*nh
        b_vals = proj_kr.get("base", {}).get(key) or [None]*np_
        for ci, val in enumerate(h_vals[:nh] + b_vals[:np_], 2):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.number_format = fmt
            cell.alignment = _align(h="right")
            if r % 2 == 0:
                cell.fill = _fill("gray")
        r += 1

    profitability_ratios = [
        ("ROE (%)", "roe_pct"), ("ROA (%)", "roa_pct"), ("PAT Margin (%)", "pat_margin_pct"),
    ]
    banking_ratios = [
        ("NIM (%)", "nim_pct"), ("GNPA (%)", "gnpa_pct"), ("NNPA (%)", "nnpa_pct"),
        ("PCR (%)", "pcr_pct"), ("Cost / Income (%)", "cost_to_income_pct"),
        ("CAR (%)", "car_pct"), ("Tier 1 (%)", "tier1_pct"),
        ("CASA Ratio (%)", "casa_ratio_pct"), ("Credit Cost (%)", "credit_cost_pct"),
    ]
    nonbank_ratios = [
        ("EBITDA Margin (%)", "ebitda_margin_pct"), ("Gross Margin (%)", "gross_margin_pct"),
        ("ROCE (%)", "roce_pct"), ("Debt / Equity (x)", "debt_to_equity"),
        ("Interest Coverage (x)", "interest_coverage"),
    ]

    merge_hdr(ws, r, 1, len(all_yrs)+1, "PROFITABILITY", bg="very_light", fg="dark_blue"); r+=1
    for lbl, key in profitability_ratios:
        write_kr_row(lbl, key)

    merge_hdr(ws, r, 1, len(all_yrs)+1,
              "BANKING-SPECIFIC RATIOS" if is_banking else "OPERATING RATIOS",
              bg="very_light", fg="dark_blue"); r+=1
    ratios_to_show = banking_ratios if is_banking else nonbank_ratios
    for lbl, key in ratios_to_show:
        write_kr_row(lbl, key)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 6: ASSUMPTIONS
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Assumptions")
    ws.freeze_panes = "B4"
    col_w(ws, [(1,35), (2,14), (3,14), (4,14), (5,14), (6,14),
               (7,14), (8,14), (9,14), (10,14), (11,14), (12,14),
               (13,14), (14,14), (15,14), (16,14)])

    r = 1
    w(ws, r, 1, f"ASSUMPTIONS  —  {company_name} ({ticker})",
      bold=True, bg="dark_blue", fg="white", size=12); r += 1

    # Macro assumptions
    macro = asmp.get("macro", {})
    w(ws, r, 1, "MACRO ASSUMPTIONS", bold=True, bg="med_blue", fg="white"); r += 1
    hdr(ws, r, ["Driver", "Base", "Bull", "Bear"], c0=1, bg="very_light", fg="dark_blue"); r += 1
    for label, key in [("RBI Repo Rate (%)", "rbi_repo_rate_pct"),
                       ("India GDP Growth (%)", "india_gdp_growth_pct"),
                       ("India CPI (%)", "india_cpi_pct")]:
        d = macro.get(key, {})
        if isinstance(d, dict):
            w(ws, r, 1, label)
            w(ws, r, 2, d.get("base"), h="center")
            w(ws, r, 3, d.get("bull"), h="center")
            w(ws, r, 4, d.get("bear"), h="center")
        r += 1
    w(ws, r, 1, "G-Sec 10yr Yield (%)")
    w(ws, r, 2, macro.get("gsec_10yr_yield_pct"), h="center"); r += 2

    # Revenue driver assumptions
    rev_drivers = asmp.get("revenue_drivers", {})
    w(ws, r, 1, "REVENUE DRIVERS", bold=True, bg="med_blue", fg="white"); r += 1
    scenario_cols = proj_yrs
    hdr_labels = ["Driver / Scenario"] + [f"{yr}\nBase" for yr in proj_yrs] + \
                 [f"{yr}\nBull" for yr in proj_yrs] + [f"{yr}\nBear" for yr in proj_yrs]
    hdr(ws, r, ["Driver"] + [f"{yr} Base" for yr in proj_yrs] +
                             [f"{yr} Bull" for yr in proj_yrs] +
                             [f"{yr} Bear" for yr in proj_yrs],
        c0=1, bg="very_light", fg="dark_blue"); r += 1

    driver_keys = (
        [("Loan Growth (%)", "loan_growth_pct"),
         ("NIM (bps)", "nim_bps"),
         ("Fee Income Growth (%)", "fee_income_growth")]
        if is_banking else
        [("Revenue Growth (%)", "revenue_growth_pct"),
         ("EBITDA Margin (%)", "ebitda_margin_pct"),
         ("Volume Growth (%)", "volume_growth_pct")]
    )
    cost_driver_keys = (
        [("Cost / Income (%)", "cost_to_income_pct"),
         ("Credit Cost (%)", "credit_cost_pct"),
         ("GNPA (%)", "gnpa_pct")]
        if is_banking else
        [("Gross Margin (%)", "gross_margin_pct"),
         ("SG&A % Revenue", "sga_pct_of_revenue"),
         ("OpEx Growth (%)", "opex_growth_pct")]
    )

    all_driver_keys = driver_keys + cost_driver_keys
    all_drivers = {**rev_drivers, **asmp.get("cost_drivers", {})}

    for lbl, key in all_driver_keys:
        d = all_drivers.get(key, {})
        w(ws, r, 1, lbl)
        for ci, yr in enumerate(proj_yrs, 2):
            yr_d = d.get(yr, {})
            if isinstance(yr_d, dict):
                w(ws, r, ci,                yr_d.get("base"), h="center")
                w(ws, r, ci + np_,          yr_d.get("bull"), h="center",
                  fg="dark_green")
                w(ws, r, ci + np_ * 2,      yr_d.get("bear"), h="center",
                  fg="dark_red")
        r += 1

    r += 1
    w(ws, r, 1, "RATIONALE", bold=True, bg="med_blue", fg="white"); r += 1
    rationale = asmp.get("rationale", {})
    for lbl, key in [("Key Assumptions", "key_assumptions"),
                     ("Bull Case Drivers", "bull_drivers"),
                     ("Bear Case Risks",  "bear_risks")]:
        w(ws, r, 1, lbl, bold=True)
        cell = ws.cell(row=r, column=2, value=rationale.get(key, ""))
        cell.font = _font()
        cell.alignment = _align(wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r+2, end_column=8)
        ws.row_dimensions[r].height = 50
        r += 3

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 7: WACC & COST OF CAPITAL
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("WACC")
    ws.sheet_view.showGridLines = False
    col_w(ws, [(1,40),(2,20),(3,15)])

    r = 1
    w(ws, r, 1, f"WACC & COST OF CAPITAL  —  {company_name} ({ticker})",
      bold=True, bg="dark_blue", fg="white", size=12); r += 2

    def wacc_row(label, value, fmt=None, bold=False, bg=None):
        nonlocal r
        w(ws, r, 1, label, bold=bold, bg=bg)
        cell = w(ws, r, 2, value, bold=bold, bg=bg, h="right")
        if fmt: cell.number_format = fmt
        r += 1

    w(ws, r, 1, "COST OF EQUITY (CAPM)",   bold=True, bg="very_light", fg="dark_blue")
    w(ws, r, 2, "Value",                    bold=True, bg="very_light", fg="dark_blue",h="right")
    r += 1
    wacc_row("Risk-Free Rate (G-Sec 10yr, %)",    wacc_d.get("risk_free_rate_pct"))
    wacc_row("+ Equity Risk Premium (%)",          wacc_d.get("equity_risk_premium_pct"))
    wacc_row("× Beta",                            wacc_d.get("beta"))
    wacc_row("= Cost of Equity (%)",               wacc_d.get("cost_of_equity_pct"), bold=True, bg="green")
    r += 1

    w(ws, r, 1, "COST OF DEBT", bold=True, bg="very_light", fg="dark_blue"); r += 1
    wacc_row("Cost of Debt (Pre-tax, %)",          wacc_d.get("cost_of_debt_pre_tax_pct"))
    wacc_row("Tax Rate (%)",                       wacc_d.get("tax_rate_pct"))
    wacc_row("= Cost of Debt (Post-tax, %)",       wacc_d.get("cost_of_debt_post_tax_pct"), bold=True)
    r += 1

    w(ws, r, 1, "WACC BUILD-UP", bold=True, bg="very_light", fg="dark_blue"); r += 1
    wacc_row("Equity Weight (%)",                  wacc_d.get("equity_weight_pct"))
    wacc_row("Debt Weight (%)",                    wacc_d.get("debt_weight_pct"))
    wacc_row("WACC (%)",                           wacc_d.get("wacc_pct"), bold=True, bg="light_blue")
    r += 1
    wacc_row("Terminal Growth Rate (%)",           wacc_d.get("terminal_growth_rate_pct"))
    r += 1

    w(ws, r, 1, "RATIONALE", bold=True, fg="dark_blue"); r += 1
    cell = ws.cell(row=r, column=1, value=wacc_d.get("wacc_rationale", ""))
    cell.font = _font(italic=True)
    cell.alignment = _align(wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=3)
    ws.row_dimensions[r].height = 80

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 8: VALUATION
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    col_w(ws, [(1,35),(2,18),(3,18),(4,18),(5,18)])

    r = 1
    w(ws, r, 1, f"VALUATION  —  {company_name} ({ticker})",
      bold=True, bg="dark_blue", fg="white", size=12); r += 2

    def val_row(label, value, fmt=FMT_CR, bold=False, fg_c="black"):
        nonlocal r
        w(ws, r, 1, label, bold=bold)
        cell = w(ws, r, 2, value, bold=bold, fg=fg_c, h="right")
        if fmt: cell.number_format = fmt
        r += 1

    # DCF
    dcf = val_d.get("dcf", {})
    w(ws, r, 1, "DCF VALUATION (FCFE Method)", bold=True, bg="med_blue", fg="white"); r += 1
    hdr(ws, r, [""] + proj_yrs, c0=1, bg="very_light", fg="dark_blue"); r += 1
    w(ws, r, 1, "FCFE (₹ Cr)")
    fcfe = dcf.get("fcfe_base_cr") or [None]*np_
    for ci, val in enumerate(fcfe[:np_], 2):
        cell = ws.cell(row=r, column=ci, value=val)
        cell.number_format = FMT_CR
        cell.alignment = _align(h="right")
    r += 2
    val_row("PV of FCFE (₹ Cr)",               dcf.get("pv_fcfe_cr"))
    val_row("Terminal Value (₹ Cr)",            dcf.get("terminal_value_cr"))
    val_row("PV of Terminal Value (₹ Cr)",      dcf.get("pv_terminal_value_cr"))
    val_row("Total Equity Value (₹ Cr)",        dcf.get("total_equity_value_cr"), bold=True)
    val_row("DCF Fair Value per Share (₹)",     dcf.get("fair_value_per_share"),
            fmt=FMT_RS, bold=True, fg_c="dark_blue")
    val_row("Weight in Blended (%)",            dcf.get("weight_pct"))
    r += 1

    # P/E Comps
    pe = val_d.get("pe_comps", {})
    w(ws, r, 1, "P/E COMPARABLE ANALYSIS", bold=True, bg="med_blue", fg="white"); r += 1
    hdr(ws, r, ["Company", "TTM P/E"], c0=1, bg="very_light", fg="dark_blue"); r += 1
    for peer in (pe.get("peers") or []):
        w(ws, r, 1, peer.get("name", ""))
        cell = ws.cell(row=r, column=2, value=peer.get("pe_ttm"))
        cell.number_format = "0.0x"; cell.alignment = _align(h="right"); r += 1
    val_row("Sector Median P/E",               pe.get("sector_median_pe"), fmt="0.0x")
    val_row(f"Target P/E ({pe.get('base_year','FY27E')} EPS: ₹{pe.get('base_year_eps',0):.0f})",
            pe.get("target_pe"), fmt="0.0x")
    val_row("P/E Fair Value per Share (₹)",    pe.get("fair_value_per_share"),
            fmt=FMT_RS, bold=True, fg_c="dark_blue")
    val_row("Weight (%)", pe.get("weight_pct")); r += 1

    # P/BV Comps
    pbv = val_d.get("pbv_comps", {})
    w(ws, r, 1, "P/BV COMPARABLE ANALYSIS", bold=True, bg="med_blue", fg="white"); r += 1
    hdr(ws, r, ["Company", "P/BV"], c0=1, bg="very_light", fg="dark_blue"); r += 1
    for peer in (pbv.get("peers") or []):
        w(ws, r, 1, peer.get("name", ""))
        cell = ws.cell(row=r, column=2, value=peer.get("pbv"))
        cell.number_format = "0.0x"; cell.alignment = _align(h="right"); r += 1
    val_row("Sector Median P/BV",              pbv.get("sector_median_pbv"), fmt="0.0x")
    val_row("Target P/BV",                     pbv.get("target_pbv"), fmt="0.0x")
    val_row("P/BV Fair Value per Share (₹)",   pbv.get("fair_value_per_share"),
            fmt=FMT_RS, bold=True, fg_c="dark_blue")
    val_row("Weight (%)", pbv.get("weight_pct")); r += 1

    # GGM
    ggm = val_d.get("ggm", {})
    w(ws, r, 1, "GORDON GROWTH MODEL", bold=True, bg="med_blue", fg="white"); r += 1
    val_row("Sustainable ROE (%)",             ggm.get("sustainable_roe_pct"))
    val_row("Terminal Growth (%)",             ggm.get("terminal_growth_pct"))
    val_row("Cost of Equity (%)",              ggm.get("cost_of_equity_pct"))
    val_row("Implied P/BV",                    ggm.get("implied_pbv"), fmt="0.00x")
    val_row("GGM Fair Value per Share (₹)",    ggm.get("fair_value_per_share"),
            fmt=FMT_RS, bold=True, fg_c="dark_blue")
    val_row("Weight (%)", ggm.get("weight_pct")); r += 1

    # Blended
    w(ws, r, 1, "BLENDED VALUATION SUMMARY", bold=True, bg="dark_blue", fg="white"); r += 1
    blended = val_d.get("blended", {})
    val_row("Blended Fair Value (₹)",          blended.get("fair_value_per_share"),
            fmt=FMT_RS, bold=True, fg_c="dark_blue")
    val_row("CMP (₹)",                         blended.get("cmp"), fmt=FMT_RS)
    val_row("Upside / (Downside) %",           blended.get("upside_pct"), fmt="0.0")
    w(ws, r, 1, "Rating", bold=True)
    w(ws, r, 2, rating, bold=True, fg=rc, h="right", size=12); r += 1
    val_row("12M Target Price (₹)",            blended.get("target_price_12m"),
            fmt=FMT_RS, bold=True, fg_c=rc); r += 1

    sc = val_d.get("scenario_targets", {})
    hdr(ws, r, ["Scenario", "Bull", "Base", "Bear"], c0=1, bg="very_light", fg="dark_blue"); r += 1
    w(ws, r, 1, "Target Price (₹)")
    for ci, key in enumerate(["bull","base","bear"], 2):
        cell = ws.cell(row=r, column=ci, value=sc.get(key))
        cell.number_format = FMT_RS
        cell.alignment = _align(h="right")
        cell.font = _font(bold=True, color=["dark_green","black","dark_red"][ci-2])
    r += 1

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 9: SENSITIVITY
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    col_w(ws, [(i, 14) for i in range(1, 10)])

    r = 1
    w(ws, r, 1, f"SENSITIVITY ANALYSIS  —  {company_name} ({ticker})",
      bold=True, bg="dark_blue", fg="white", size=12); r += 2

    def write_sens_grid(ws, start_row, grid_data, title):
        r = start_row
        w(ws, r, 1, title, bold=True, bg="med_blue", fg="white"); r += 1
        row_label = grid_data.get("row_label", "")
        col_label = grid_data.get("col_label", "")
        row_vals  = grid_data.get("row_values", [])
        col_vals  = grid_data.get("col_values", [])
        grid      = grid_data.get("grid", [])

        # Corner label
        w(ws, r, 1, f"{row_label} \\ {col_label}", bold=True, bg="very_light",
          fg="dark_blue", wrap=True)
        for ci, cv in enumerate(col_vals, 2):
            w(ws, r, ci, cv, bold=True, bg="very_light", fg="dark_blue", h="center")
        r += 1

        base_val = blend.get("fair_value_per_share") or 1
        for ri, (rv, row) in enumerate(zip(row_vals, grid)):
            w(ws, r, 1, rv, bold=True, bg="very_light", fg="dark_blue", h="center")
            for ci, cell_val in enumerate(row, 2):
                cell = ws.cell(row=r, column=ci, value=cell_val)
                cell.number_format = FMT_CR0
                cell.alignment = _align(h="right")
                # Color: green if > base, red if < base
                if cell_val and base_val:
                    try:
                        ratio = float(cell_val) / float(base_val)
                        if ratio >= 1.10:
                            cell.fill = _fill("green")
                        elif ratio <= 0.90:
                            cell.fill = _fill("red_light")
                    except Exception:
                        pass
            r += 1
        return r + 1

    r = write_sens_grid(ws, r, sens_d.get("dcf_grid", {}),
                        "DCF SENSITIVITY: Fair Value vs WACC & Terminal Growth Rate")
    r = write_sens_grid(ws, r, sens_d.get("pe_grid", {}),
                        "P/E SENSITIVITY: Fair Value vs EPS CAGR & Target P/E")
    r = write_sens_grid(ws, r, sens_d.get("pbv_grid", {}),
                        "P/BV SENSITIVITY: Fair Value vs ROE & Target P/BV")

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 10: SCENARIO SUMMARY
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Scenario Summary")
    ws.sheet_view.showGridLines = False
    col_w(ws, [(1,35),(2,18),(3,18),(4,18)])

    r = 1
    w(ws, r, 1, f"SCENARIO SUMMARY  —  {company_name} ({ticker})",
      bold=True, bg="dark_blue", fg="white", size=12); r += 2

    hdr(ws, r, ["Metric", "Bull Case", "Base Case", "Bear Case"],
        c0=1, bg="med_blue"); r += 1

    def sc_row(label, bull_v, base_v, bear_v, fmt=FMT_CR, bold=False):
        nonlocal r
        w(ws, r, 1, label, bold=bold)
        for ci, (val, color) in enumerate(
            [(bull_v,"dark_green"),(base_v,"black"),(bear_v,"dark_red")], 2):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.number_format = fmt
            cell.alignment = _align(h="right")
            cell.font = _font(bold=bold, color=color)
        r += 1

    def sc_val(scenario, key, idx):
        lst = proj_is.get(scenario, {}).get(key) or [None]*np_
        return lst[idx] if idx < len(lst) else None

    def sc_kr(scenario, key, idx):
        lst = proj_kr.get(scenario, {}).get(key) or [None]*np_
        return lst[idx] if idx < len(lst) else None

    sc_row("Target Price (₹)",
           sc.get("bull"), sc.get("base"), sc.get("bear"), fmt=FMT_RS, bold=True)
    sc_row("Upside / (Downside) %", None, blended.get("upside_pct"), None, fmt="0.0")
    r += 1

    merge_hdr(ws, r, 1, 4, "FY27E KEY METRICS", bg="very_light", fg="dark_blue"); r += 1
    idx27 = 1  # FY27E is index 1 in proj_yrs
    sc_row("PAT (₹ Cr)",
           sc_val("bull","pat",idx27), sc_val("base","pat",idx27), sc_val("bear","pat",idx27))
    sc_row("EPS (₹)",
           sc_val("bull","eps_basic",idx27), sc_val("base","eps_basic",idx27),
           sc_val("bear","eps_basic",idx27), fmt=FMT_EPS)
    if is_banking:
        sc_row("NIM (%)",
               sc_kr("bull","nim_pct",idx27), sc_kr("base","nim_pct",idx27),
               sc_kr("bear","nim_pct",idx27), fmt=FMT_PCT)
        sc_row("GNPA (%)",
               sc_kr("bull","gnpa_pct",idx27), sc_kr("base","gnpa_pct",idx27),
               sc_kr("bear","gnpa_pct",idx27), fmt=FMT_PCT)
    else:
        sc_row("EBITDA Margin (%)",
               sc_kr("bull","ebitda_margin_pct",idx27), sc_kr("base","ebitda_margin_pct",idx27),
               sc_kr("bear","ebitda_margin_pct",idx27), fmt=FMT_PCT)
    sc_row("ROE (%)",
           sc_kr("bull","roe_pct",idx27), sc_kr("base","roe_pct",idx27),
           sc_kr("bear","roe_pct",idx27), fmt=FMT_PCT)

    r += 2
    merge_hdr(ws, r, 1, 4, "5-YEAR CAGR SUMMARY (FY25→FY30E)", bg="very_light", fg="dark_blue"); r += 1
    cagr = proj.get("cagr_summary", {})
    for lbl, key in [("Historical PAT CAGR (FY22→FY25)", "pat_fy22_fy25_pct"),
                     ("Projected PAT CAGR (FY25→FY30E, Base)", "pat_fy25_fy30_base_pct"),
                     ("Projected EPS CAGR (FY25→FY30E, Base)", "eps_fy25_fy30_base_pct"),
                     ("Projected BVPS CAGR (FY25→FY30E, Base)", "bvps_fy25_fy30_base_pct")]:
        w(ws, r, 1, lbl)
        cell = ws.cell(row=r, column=2, value=cagr.get(key))
        cell.number_format = "0.0"
        cell.alignment = _align(h="right")
        r += 1

    # ── Save ────────────────────────────────────────────────────────────────
    # Verify balance sheet before saving
    h_assets = hist_bs.get("total_assets") or []
    h_liab   = hist_bs.get("total_liabilities") or []
    h_eq     = hist_bs.get("total_equity") or []
    for i, yr in enumerate(hist_yrs):
        if i < len(h_assets) and i < len(h_liab) and i < len(h_eq):
            if all(x is not None for x in [h_assets[i], h_liab[i], h_eq[i]]):
                diff = abs(h_assets[i] - h_liab[i] - h_eq[i])
                if diff > 5:
                    print(f"  ⚠️  BS imbalance {yr}: diff = {diff:.0f} Cr")

    wb.save(out_path)
    size_kb = os.path.getsize(out_path) / 1024
    print(f"\n✅  Excel saved → {out_path}  ({size_kb:.0f} KB)")
    print(f"    Sheets: {[s.title for s in wb.worksheets]}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════

def generate_financial_model(company_name: str, ticker: str,
                              exchange: str = "NSE",
                              sector: str   = "Banking / NBFC",
                              out_dir: str  = "/tmp/financial_models"):

    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{ticker}_model.xlsx")

    print(f"\n{'='*60}")
    print(f"  Financial Model v4  |  {company_name} ({ticker})")
    print(f"  Exchange: {exchange}   Sector: {sector}")
    print(f"  Architecture: 2 LLM calls + deterministic Excel builder")
    print(f"{'='*60}\n")

    # ── Pre-fetch free data ──────────────────────────────────────────────
    print("📡 Fetching yfinance data ...")
    yf_data = fetch_yfinance_data(f"{ticker}.NS")

    time.sleep(1)

    print("📡 Scraping screener.in ...")
    screener_text = fetch_screener_data(ticker)
    print(f"   screener.in: {len(screener_text)} chars")

    # ── Call 1 ──────────────────────────────────────────────────────────
    model_json = run_analysis(company_name, ticker, exchange, sector,
                              yf_data, screener_text)

    # Save JSON for debugging
    json_path = os.path.join(out_dir, f"{ticker}_model.json")
    with open(json_path, "w") as f:
        json.dump(model_json, f, indent=2, default=str)
    print(f"   JSON saved → {json_path}")

    # ── Call 2 ──────────────────────────────────────────────────────────
    thesis_json = run_thesis(company_name, ticker, model_json)

    # ── Build Excel ──────────────────────────────────────────────────────
    print("📊 Building Excel model ...")
    try:
        build_excel(model_json, thesis_json, out_path, company_name, ticker)
    except Exception as e:
        traceback.print_exc()
        print(f"\n❌  Excel build failed: {e}")
        raise

    return out_path


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    generate_financial_model(
        company_name = "HDFC Bank Ltd",
        ticker       = "HDFCBANK",
        exchange     = "NSE",
        sector       = "Banking / NBFC",
        out_dir      = "/tmp/financial_models",
    )
